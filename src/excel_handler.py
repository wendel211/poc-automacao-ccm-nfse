from __future__ import annotations
from datetime import datetime
from pathlib import Path
from typing import Iterator

import pandas as pd
from loguru import logger
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from src.models import InputRow, Municipio, RowResult, StatusExecucao

_COL_MAP = {
    "ID do documento": "id_documento",
    "Empresa": "empresa",
    "Nº documento": "num_documento",
    "Fornecedor": "fornecedor",
    "Nome fornecedor": "nome_fornecedor",
    "CNPJ": "cnpj_raw",
    "CCM": "ccm_existente",
    "Referência": "referencia",
    "MUNICIPIO": "municipio",
    "COD.VERIFICACAO": "cod_verificacao",
}

_OUTPUT_COLS = [
    "STATUS_EXECUCAO",
    "MENSAGEM_TECNICA",
    "CCM_ENCONTRADO",
    "RAZAO_SOCIAL",
    "SITUACAO_CADASTRAL",
    "ATIVIDADE_PRINCIPAL",
    "FONTE_CADASTRO",
    "TIPO_DOCUMENTO",
    "NOTA_MUNICIPIO",
    "NOTA_NUMERO",
    "NOTA_COMPETENCIA",
    "CNPJ_EMITENTE_CONFERE",
    "CHAVE_DV_VALIDO",
    "ARQUIVO_CADASTRO",
    "ARQUIVO_CADASTRO_PUBLICO",
    "ARQUIVO_DADOS_NOTA",
    "ARQUIVO_NOTA_PDF",
    "ARQUIVO_NOTA_XML",
    "ARQUIVO_EVIDENCIA_CADASTRO",
    "ARQUIVO_EVIDENCIA_NOTA",
    "MUNICIPIO_ESTRATEGIA",
    "DATA_EXECUCAO",
]

_STATUS_COLORS = {
    StatusExecucao.SUCESSO: "C6EFCE",
    StatusExecucao.PARCIAL: "FFEB9C",
    StatusExecucao.ERRO: "FFC7CE",
    StatusExecucao.INDISPONIVEL: "D9D9D9",
}


def read_rows(path: Path) -> Iterator[InputRow]:
    df = pd.read_excel(path, dtype=str).fillna("")
    for _, row in df.iterrows():
        raw = {v: str(row.get(k, "")).strip() for k, v in _COL_MAP.items()}
        try:
            mun_key = raw["municipio"].upper().strip()
            raw["municipio"] = Municipio(mun_key)
            if not raw["cnpj_raw"]:
                logger.warning("Linha sem CNPJ ignorada: id={}", raw.get("id_documento"))
                continue
            if not raw["cod_verificacao"]:
                logger.warning("Linha sem COD.VERIFICACAO ignorada: id={}", raw.get("id_documento"))
                continue
            yield InputRow(**raw)
        except Exception as exc:
            logger.error("Linha inválida ignorada: {} — {}", raw.get("id_documento"), exc)


def write_results(source_path: Path, dest_path: Path, results: dict[str, RowResult]) -> None:
    wb = load_workbook(source_path)
    ws = wb.active

    header_row = [cell.value for cell in ws[1]]

    for col_name in _OUTPUT_COLS:
        if col_name not in header_row:
            ws.cell(row=1, column=len(header_row) + 1, value=col_name).font = Font(bold=True)
            header_row.append(col_name)

    id_col_idx = header_row.index("ID do documento") + 1

    for row_idx in range(2, ws.max_row + 1):
        doc_id = str(ws.cell(row=row_idx, column=id_col_idx).value or "").strip()
        result = results.get(doc_id)
        if not result:
            continue

        fill = PatternFill("solid", start_color=_STATUS_COLORS.get(result.status, "FFFFFF"))
        values = [
            result.status.value,
            result.mensagem_tecnica or "",
            result.ccm_encontrado or "",
            result.razao_social or "",
            result.situacao_cadastral or "",
            result.atividade_principal or "",
            result.fonte_cadastro or "",
            result.tipo_documento or "",
            result.nota_municipio or "",
            result.nota_numero or "",
            result.nota_competencia or "",
            result.cnpj_emitente_confere or "",
            result.chave_dv_valido or "",
            result.arquivo_cadastro or "",
            result.arquivo_cadastro_publico or "",
            result.arquivo_dados_nota or "",
            result.arquivo_nota_pdf or "",
            result.arquivo_nota_xml or "",
            result.arquivo_evidencia_cadastro or "",
            result.arquivo_evidencia_nota or result.arquivo_evidencia or "",
            result.municipio_estrategia or "",
            result.data_execucao or datetime.now().isoformat(timespec="seconds"),
        ]
        for col_name, value in zip(_OUTPUT_COLS, values):
            col_idx = header_row.index(col_name) + 1
            # Normalize Unicode punctuation to ASCII so openpyxl does not emit
            # characters that Windows cp1252 Excel installations misread.
            safe_value = str(value).replace("—", " - ").replace("–", " - ")
            cell = ws.cell(row=row_idx, column=col_idx, value=safe_value)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=False)

        # Requisito do enunciado: "Caso encontrado, gravar valor em CCM".
        # Preenche a coluna ORIGINAL `CCM` (vinha vazia na planilha de entrada)
        # com a inscricao municipal encontrada, mantendo tambem CCM_ENCONTRADO.
        if result.ccm_encontrado and "CCM" in header_row:
            ccm_idx = header_row.index("CCM") + 1
            ccm_cell = ws.cell(row=row_idx, column=ccm_idx, value=str(result.ccm_encontrado))
            ccm_cell.fill = fill

    _autosize_output_columns(ws, header_row)
    wb.save(dest_path)
    logger.info("Planilha de saida salva: {}", dest_path)


def _autosize_output_columns(ws, header_row: list) -> None:
    for col_name in _OUTPUT_COLS:
        if col_name not in header_row:
            continue
        col_idx = header_row.index(col_name) + 1
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        max_len = len(col_name)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2):
            cell_val = str(row[0].value or "")
            max_len = max(max_len, len(cell_val))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)
